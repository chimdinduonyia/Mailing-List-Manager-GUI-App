from tkinter import *
from tkinter import filedialog as fd
from openpyxl import Workbook
from openpyxl import load_workbook
from tkinter import messagebox
from time import sleep
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib
import mimetypes
import email.mime.application
import re

root = Tk()
root.title("Newsletter App")
root.geometry("700x700")

# Globals: mailing list filename, active workbook, active worksheet, dictionary of attachments, email username & password, ready_to_send boolean
xl_filename = ""
active_sheet = None
workbook = None
attachments = {}
email_user = ""
email_pass = ""
ready = False

# -----------------------------------------------------------------------------Functions

# This function gets the e-mail addresses in a mailing list and casts them to a list of recipients
def get_mails(work_sheet):
    recipients = []
    for column in work_sheet.iter_cols(min_row=2, min_col=5, values_only=True):
        for address in column:
            recipient = address
            recipients.append(recipient)
    # This joins all the addresses with a comma and casts them to a cc variable that will be the recipient field of the e-mail
    cc = ",".join(recipients)

    # Accept the subject and body values
    subject = subject_field.get()
    body = body_field.get("1.0", END)

    return cc, subject, body

#Function to populate the recipients section
def update_recipients():

    global active_sheet
    global recipient_field

    to, subject, body = get_mails(active_sheet)
    recipient_field.delete(0, END)
    recipient_field.insert(0, to)


# New File Button Function
def new_file():
    global xl_filename
    global new_file
    global active_sheet
    global workbook
    global ml_label
    # Create a new workbook and adds default fields
    workbook = Workbook()
    active_sheet = workbook.active
    active_sheet["A1"].value = "S/N"
    active_sheet["B1"].value = "Surname"
    active_sheet["C1"].value = "Firstname"
    active_sheet["D1"].value = "Phone"
    active_sheet["E1"].value = "E-mail"
    # Provides dialog and casts selections to a tuple which is stored as the variable filetypes
    filetypes = (
        ("Excel Spreadsheet", "*.xlsx"),
    )
    xl_filename = fd.asksaveasfilename(
        initialfile="Untitled.xlsx", initialdir="./Mailing Lists", defaultextension=".xlsx", filetypes=filetypes)
    # Check if nothing was returned i.e user chose to cancel and set filename back to default ""
    if xl_filename == None:
        xl_filename = ""
    else:
        # If not, save the workbook
        workbook.save(filename=xl_filename)
        # display name of opened spreadsheet at top of form: uses regex to isolate the name of the file from the filepath
        current_list = re.search(r"/([^/]+)$", xl_filename).group(1)
        # Mailing list label at the top of the form shows the name of the active mailing list
        ml_label = Label(recipient_form, text="Now Open: " +
                         current_list, anchor=CENTER)
        ml_label.grid(row=0, column=0, columnspan=2, sticky=W+E)

# Generator function that keeps track of row s/n: selects first column (S/N) and yields all the integers in that column
def track_row():
    global active_sheet
    
    for row in active_sheet.iter_rows(min_row=2, values_only=True, max_col=1):
        yield row[0]

# Function to execute when the save button is clicked
def save_row(work_sheet):
    # Variables
    global workbook
    global sur_field
    global fir_field
    global phone_field
    global mail_field
    global save_label

    # Get input at Entries and Validate that they are formatted properly
    surname = sur_field.get()
    firstname = fir_field.get()
    phone_number = phone_field.get()
    email = mail_field.get()
    # Check if all the input fields have been populated
    if surname != "" and firstname != "" and phone_number != "" and email != "":
        mail_pattern = r'^[a-z0-9]+[\._]?[ a-z0-9]+[@]\w+[. ]\w{2,3}$'
        # Check if mail address entered is formatted as a valid e-mail address
        if re.search(mail_pattern, email):
            # Check if worksheet is an empty worksheet or if next cell is an empty cell
            if work_sheet["A2"].value:
                # If not, check to see the S/N of the last row, What is the highest row number in the spreadsheet?
                last_row = max(track_row())
                new_row = str(int(last_row) + 2)
                # Fix values in the rows after that
                # Enter row number
                work_sheet["A" + new_row] = str(int(last_row) + 1)
                work_sheet["B" + new_row] = surname
                work_sheet["C" + new_row] = firstname
                work_sheet["D" + new_row] = phone_number
                work_sheet["E" + new_row] = email
                saveit = True
            else:
                # If so, fix the values in the cells
                work_sheet["A2"] = '1'
                work_sheet["B2"] = surname
                work_sheet["C2"] = firstname
                work_sheet["D2"] = phone_number
                work_sheet["E2"] = email
                saveit = True
        else:
            # Show error message for badly formatted e-mail
            messagebox.showinfo(
                "Invalid Entry", "The E-mail you entered is not a valid e-mail")
            saveit = False
    else:
        # Show error message for any missing fields
        messagebox.showinfo("Incomplete Entries", "All fields are required")
        saveit = False
    # save workbook
    if saveit:
        # If saveit == True, save the entries
        workbook.save(filename=xl_filename)
        sur_field.delete(0, END)
        fir_field.delete(0, END)
        phone_field.delete(0, END)
        mail_field.delete(0, END)
        # Add a Label that displays when something has been saved
        save_label = Label(recipient_form, text="Saved!")
        save_label.grid(row=5, column=0, columnspan=2, sticky=W+E)
        recipient_form.after(2000, lambda: save_label.destroy())

    update_recipients()

# Function to open an already created work_book
def open_file():
    global xl_filename
    global active_sheet
    global workbook
    global ml_label
    filetypes = (
        ("Excel Spreadsheet", "*.xlsx"),
        ("Excel Spreadsheet", "*.xlsm")
    )
    # Provides dialog and casts selections to a tuple which is stored as the variable filename
    xl_filename = fd.askopenfilename(
        initialfile="Untitled.xlsx", initialdir="C://Users/USER/Documents/PYTHON PROJECTS/Mailing Lists", defaultextension=".xlsx", filetypes=filetypes)
    # If User presses cancel on dialog, revert back to default global filename
    if xl_filename == None:
        xl_filename = ""
    # If User enters filename, make it active and print name of file at top, excluding the rest of the path
    if len(xl_filename) > 1:
        workbook = load_workbook(filename=xl_filename)
        active_sheet = workbook.active
        current_list = re.search(r"/([^/]+)$", xl_filename).group(1)
        ml_label = Label(recipient_form, text="Now Open: " +
                         current_list, anchor=CENTER)
        ml_label.grid(row=0, column=0, columnspan=2, sticky=W+E)
    else:
        pass

    update_recipients()

# -----------------------------------------------------------------E-mailing functionality



def get_id_pass():
    global user_id
    global user_pass
    global user_label
    global pass_label
    global submit_btn

    # Creates a new frame for the username and pass to be entered    login_frame = Frame(root)
    login_frame.grid(row=1, column=1, sticky=W+E, pady=20, padx=10)
    user_label = Label(login_frame, text="Username: ")
    user_label.grid(row=0, column=0, sticky="W", padx=5, pady=5)
    user_id = Entry(login_frame, bd=3)
    user_id.grid(row=0, column=1, sticky="W", padx=5, pady=5)
    pass_label = Label(login_frame, text="Password: ")
    pass_label.grid(row=1, column=0, sticky="W", padx=5, pady=5)
    user_pass = Entry(login_frame, show="*", bd=3)
    user_pass.grid(row=1, column=1, sticky="W", padx=5, pady=5)
    submit_btn = Button(login_frame, text="Submit",
                        bg="#2e6930", fg="#fff", command=submit)
    submit_btn.grid(row=2, column=1, sticky="W", padx=5, pady=5)



def submit():
    global user_id
    global user_pass
    global user_label
    global pass_label
    global attachments
    global submit_btn
    global subject_field
    global body_field
    global frame_2
    global recipient_field

    email_user = user_id.get()
    email_pass = user_pass.get()
    print("Gotten Credentials")
    user_id.grid_forget()
    user_pass.grid_forget()
    user_label.grid_forget()
    pass_label.grid_forget()
    submit_btn.grid_forget()
    ready = True
    if ready:
        print("Ready to connect")
        # initialises connection to smtp server
        smtp_ssl_host = 'smtp.gmail.com'
        smtp_ssl_port = 465
        s = smtplib.SMTP_SSL(smtp_ssl_host, smtp_ssl_port)
        try:
            s.login(email_user, email_pass)
            print("Login Successful")
            connect = True
            print("Connected")
        except smtplib.SMTPAuthenticationError:
            messagebox.showinfo("Login Unsuccessful")
            pass
        while connect:
            # Get the recipients, the subject and body of mail
            to, subject, body = get_mails(active_sheet)
            send.config(state=DISABLED)
            # instantiate message as multipartite message
            msg = MIMEMultipart()
            msg['Subject'] = subject
            msg['From'] = email_user  # email_user
            msg['To'] = recipient_field.get()
            # Add body text to mail
            txt = MIMEText(body)
            msg.attach(txt)
            for name, file in attachments.items():
                filename = file.name
                with open(filename, 'rb') as f:
                    attachment = email.mime.application.MIMEApplication(
                        f.read(), _subtype=file.name.split('.')[-1])
                    attachment.add_header(
                        'Content-Disposition', 'attachment', filename=name)
                    msg.attach(attachment)
            s.send_message(msg)
            s.quit()
            print("Mail Sent!")
            connect = False
        else:
            send.config(state=NORMAL)

    attachments = {}
    subject_field.delete(0, END)
    body_field.delete("1.0", END)
    frame_2.grid_forget()


# Function to send mail
def send_mail():
    # Display Message Box that says are you sure you want to send this mail?
    answer = messagebox.askyesno(
        title='Confirmation', message='Are you sure that you want to send?')
    if answer:
        # grids entry boxes for username and password
        # email_user
        # email_pass
        get_id_pass()

# Function for the Attach Files Button


def attach_files():
    global frame_2
    global attachments

    filetypes = (("all files", "*.*"), ("png files", "*.png"),
                 ("pdf files", "*.pdf"), ("jpeg files", "*.jpg"), ("excel spreadsheets", "*.xlsx"))
    files = fd.askopenfiles(filetypes=filetypes)
    for file in files:
        label_text = re.search(r"/([^/]+)$", file.name).group(1)
        attachments[label_text] = file
    print(attachments)
    counter = 0
    for file_name in attachments.keys():
        Label(frame_2, text=file_name, relief=SUNKEN).grid(
            row=0, column=counter, padx=5, ipadx=5, pady=10, ipady=5, sticky=W)
        counter += 1


# ---------------------------------------------------------------Recipient Form
recipient_form = LabelFrame(root, text="Recipient Form")
recipient_form.grid(row=0, column=0, padx=10, pady=10)
# Spreadsheet Name Label
ml_label = Label(recipient_form, text="Now Open: -----", anchor=CENTER)
ml_label.grid(row=0, column=0, columnspan=2, sticky=W+E)
# First Field (Surname)
sur_label = Label(recipient_form, text="Surname: ")
sur_label.grid(row=1, column=0, padx=5, pady=10, sticky="W")
sur_field = Entry(recipient_form, bd=3, width=20)
sur_field.grid(row=1, column=1, padx=5, pady=10)
# Second Field (First Name)
fir_label = Label(recipient_form, text="Firstname: ")
fir_label.grid(row=2, column=0, padx=5, pady=10, sticky="W")
fir_field = Entry(recipient_form, bd=3, width=20)
fir_field.grid(row=2, column=1, padx=5, pady=10)
# Third Field (Phone)
phone_label = Label(recipient_form, text="Phone: ")
phone_label.grid(row=3, column=0, padx=5, pady=10, sticky="W")
phone_field = Entry(recipient_form, bd=3, width=20)
phone_field.grid(row=3, column=1, padx=5, pady=10)
# Fourth Field (E-mail)
mail_label = Label(recipient_form, text="E-mail: ")
mail_label.grid(row=4, column=0, padx=5, pady=10, sticky="W")
mail_field = Entry(recipient_form, bd=3, width=20)
mail_field.grid(row=4, column=1, padx=5, pady=(10, 20))

# New, Open and Save Button
# Frame for New, Open and Save Buttons
button_frame = Frame(recipient_form)
button_frame.grid(row=6, column=0, sticky=W+E, columnspan=2)
# New Button
new_button = Button(button_frame, text="New", bg="#2e6930",
                    fg="#fff", command=new_file)
new_button.grid(row=0, column=0, padx=5, ipadx=10, pady=5, ipady=10, sticky=W)
# Open Button
open_button = Button(button_frame, text="Open",
                     bg="#2e6930", fg="#fff", command=open_file)
open_button.grid(row=0, column=1, padx=5, ipadx=10, pady=5, ipady=10, sticky=W)
# Save Button
save_button = Button(button_frame, text="Save",
                     bg="#2e6930", fg="#fff", command=lambda: save_row(active_sheet))
save_button.grid(row=0, column=2, padx=5, ipadx=10, pady=5, ipady=10, sticky=W)


# -------------------------------------------------------------------------------Compose Frame
compose_frame = LabelFrame(root, text="Compose")
compose_frame.grid(row=0, column=1, padx=10, pady=10)
# Frames inside the Compose Frame
frame_1 = Frame(compose_frame, relief=SUNKEN)
frame_1.grid(row=0, column=0, padx=5, pady=5)
frame_2 = LabelFrame(compose_frame, text="Attachments")
frame_2.grid(row=1, column=0, padx=10, sticky=W+E)
frame_3 = Frame(compose_frame)
frame_3.grid(row=2, column=0, padx=10, sticky=W+E)
# Recipients Field
recipient_label = Label(frame_1, text="Recipient(s): ")
recipient_label.grid(row=1, column=0, padx=5, sticky="W")
recipient_field = Entry(frame_1, bd=3, width=50)
recipient_field.grid(row=2, column=0, padx=5, sticky=W+E, pady=(5, 20))
# Subject Field
subject_label = Label(frame_1, text="Subject: ")
subject_label.grid(row=3, column=0, padx=5, sticky="W")
subject_field = Entry(frame_1, bd=3, width=50)
subject_field.grid(row=4, column=0, padx=5, sticky=W+E, pady=(5, 20))
# Body Field
body_label = Label(frame_1, text="Body: ")
body_label.grid(row=5, column=0, padx=5, sticky="W")
body_field = Text(frame_1, bd=3, width=50, height=10)
body_field.grid(row=6, column=0, padx=5, sticky=W+E, pady=(5, 0))
# Attach and Send Buttons
attach = Button(frame_3, text="Attach File(s)", bg="#2e6930",
                fg="#fff", command=attach_files)
attach.grid(row=0, column=0, padx=(0, 10),
            ipadx=10, pady=10, ipady=10, sticky="W")
send = Button(frame_3, text="Send -->", bg="#2e6930",
              fg="#fff", command=send_mail)
send.grid(row=0, column=1, padx=(0, 10),
          ipadx=10, pady=10, ipady=10, sticky="W")
root.mainloop()
